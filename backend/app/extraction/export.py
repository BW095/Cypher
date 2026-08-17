"""
Export helpers — convert a list of StructuredRecords to CSV, Excel, or JSON bytes.

All functions return ``(bytes_content, media_type, filename)`` so the
FastAPI route can pass them directly to a ``Response``.

Dependencies (already in requirements.txt):
  - csv (stdlib)
  - io  (stdlib)
  - json (stdlib)
  - openpyxl — for Excel export
"""

from __future__ import annotations

import csv
import io
import json
from typing import List, Tuple

from app.extraction.structured_extractor import StructuredRecord


def _flatten_record(rec: StructuredRecord) -> dict:
    """Convert a StructuredRecord to a flat string-valued dict for tabular export."""
    flat: dict = {
        "file_name": rec.file_name,
        "file_path": rec.file_path,
        "doc_domain": rec.doc_domain,
        "doc_category": rec.doc_category,
    }
    for key, val in rec.fields.items():
        if isinstance(val, list):
            flat[key] = "; ".join(str(v) for v in val) if val else ""
        elif val is None:
            flat[key] = ""
        else:
            flat[key] = str(val)
    return flat


def _all_keys(records: List[StructuredRecord]) -> List[str]:
    """Collect the union of all field keys across records (preserving insertion order)."""
    seen: dict = {}
    base = ["file_name", "file_path", "doc_domain", "doc_category"]
    for k in base:
        seen[k] = None
    for rec in records:
        for k in rec.fields:
            seen[k] = None
    return list(seen.keys())


# ---------------------------------------------------------------------------
# JSON export
# ---------------------------------------------------------------------------

def to_json(records: List[StructuredRecord]) -> Tuple[bytes, str, str]:
    """Serialize records to a pretty-printed JSON array."""
    payload = []
    for rec in records:
        payload.append({
            "file_name": rec.file_name,
            "file_path": rec.file_path,
            "doc_domain": rec.doc_domain,
            "doc_category": rec.doc_category,
            "fields": rec.fields,
        })
    content = json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")
    return content, "application/json", "extracted_fields.json"


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

def to_csv(records: List[StructuredRecord]) -> Tuple[bytes, str, str]:
    """Serialize records to UTF-8 CSV with a BOM for Excel compatibility."""
    if not records:
        return b"\xef\xbb\xbf", "text/csv; charset=utf-8", "extracted_fields.csv"

    keys = _all_keys(records)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=keys, extrasaction="ignore",
                             lineterminator="\r\n")
    writer.writeheader()
    for rec in records:
        writer.writerow(_flatten_record(rec))

    # Prepend UTF-8 BOM so Excel opens the file correctly without an import wizard.
    content = "\xef\xbb\xbf".encode("latin-1") + buf.getvalue().encode("utf-8")
    return content, "text/csv; charset=utf-8", "extracted_fields.csv"


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def to_excel(records: List[StructuredRecord]) -> Tuple[bytes, str, str]:
    """Serialize records to an .xlsx workbook using openpyxl.

    Produces one sheet ("Extracted Fields") with:
      - Bold, auto-width headers
      - Alternating row fill for readability
      - One sheet per document category (when more than one category is present)
        so invoice fields and contract fields don't share the same columns.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    # Group records by category so each category gets its own sheet.
    by_category: dict[str, list] = {}
    for rec in records:
        by_category.setdefault(rec.doc_category, []).append(rec)

    HEADER_FILL  = PatternFill("solid", fgColor="1F2D4E")
    ALT_FILL     = PatternFill("solid", fgColor="F0F4FA")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    BODY_FONT    = Font(name="Calibri", size=10)

    for category, cat_records in by_category.items():
        sheet_name = category.title()[:31]   # Excel sheet names ≤31 chars
        ws = wb.create_sheet(title=sheet_name)

        keys = _all_keys(cat_records)
        # Human-readable headers: replace underscores with spaces, title-case.
        headers = [k.replace("_", " ").title() for k in keys]

        # Write header row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        ws.row_dimensions[1].height = 20

        # Write data rows
        for row_idx, rec in enumerate(cat_records, start=2):
            flat = _flatten_record(rec)
            fill = ALT_FILL if row_idx % 2 == 0 else None
            for col_idx, key in enumerate(keys, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=flat.get(key, ""))
                cell.font = BODY_FONT
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Auto-width columns (sample first 40 rows for speed).
        for col_idx, key in enumerate(keys, start=1):
            col_letter = get_column_letter(col_idx)
            sample_vals = [headers[col_idx - 1]] + [
                _flatten_record(r).get(key, "") for r in cat_records[:40]
            ]
            max_len = max((len(str(v)) for v in sample_vals if v), default=10)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        # Freeze header row
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    content = buf.read()
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return content, media_type, "extracted_fields.xlsx"
